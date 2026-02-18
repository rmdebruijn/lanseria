#!/usr/bin/env python3
"""Extract and normalize TWX Sales Pipeline from Excel into SQLite + JSON.

Re-runnable, idempotent extraction with change detection.

Run:  python3 scripts/extract_twx_pipeline.py [--force]

Inputs:  ../../../context/Current TWX Sales Pipeline (1).xlsx
Outputs: data/twx_pipeline.db   (SQLite with 4 tables)
         config/twx_pipeline.json (pre-computed JSON for Streamlit)
"""
import argparse
import datetime
import hashlib
import json
import os
import re
import sqlite3
import sys

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MODEL_DIR = os.path.dirname(SCRIPT_DIR)
PROJECT_ROOT = os.path.dirname(os.path.dirname(MODEL_DIR))
EXCEL_PATH = os.path.join(PROJECT_ROOT, "context", "Current TWX Sales Pipeline (1).xlsx")
DB_DIR = os.path.join(MODEL_DIR, "data")
DB_PATH = os.path.join(DB_DIR, "twx_pipeline.db")
JSON_PATH = os.path.join(MODEL_DIR, "config", "twx_pipeline.json")

# ── Pipeline stages (ordered) ─────────────────────────────────────────
STAGES = [
    ("prospect", "Prospect", 1),
    ("qualified", "Qualified", 2),
    ("costed", "Costed", 3),
    ("committed", "Committed", 4),
    ("in_progress", "In Progress", 5),
    ("completed", "Completed", 6),
    ("on_hold", "On Hold", 7),
]

# ── Expected column headers (canonical) ───────────────────────────────
EXPECTED_HEADERS = [
    "Project Category",
    "Updated",
    "Project Name",
    "Area",
    "Total Project Value",
    "TWX Project Value",
    "Projected GP @30%",
    "Construction Company",
    "Client",
    "Client Funds if known",
    "Notes",
    "Risks",
    "Estimated Takeoff date",
    "Estimate Completion Date",
]


def file_checksum(path: str) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def project_id(name: str) -> str:
    return hashlib.sha256(name.strip().lower().encode()).hexdigest()[:12]


def normalize_date(val) -> str | None:
    """Convert datetime objects and date strings to ISO 8601, or None."""
    if val is None:
        return None
    if isinstance(val, datetime.datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s or s.lower() in ("unknown", "uknown", "timeframe unknown"):
        return None
    # Try dd/mm/yyyy and dd/m/yyyy
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def normalize_value(val) -> float | None:
    """Convert numeric or string value to rounded float, or None."""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if not s or s.lower() in ("unknown", "uknown", "na", "n/a"):
            return None
        try:
            return round(float(s), 2)
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return round(float(val), 2)
    return None


def normalize_string(val) -> str | None:
    """Strip whitespace, return None for empty/unknown."""
    if val is None:
        return None
    s = str(val).strip()
    if not s or s.lower() in ("unknown", "uknown"):
        return None
    if s.lower() in ("na", "n/a"):
        return None
    return s


def parse_duration_months(raw: str | None) -> float | None:
    """Extract duration in months from free-text like '6 Months', '8 weeks', '24 months but...'"""
    if raw is None:
        return None
    s = str(raw).strip().lower()
    if not s or s in ("unknown", "uknown"):
        return None

    # Check for actual date (datetime was already converted to ISO string)
    if re.match(r"\d{4}-\d{2}-\d{2}", s):
        return None

    m = re.search(r"(\d+)\s*(month|months)", s)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+)\s*(week|weeks)", s)
    if m:
        return round(float(m.group(1)) / 4.33, 1)
    return None


def infer_stage(name: str, notes: str | None, risks: str | None) -> str:
    """Rule-based pipeline stage inference from project notes and risks."""
    text = f"{notes or ''} {risks or ''}".lower()
    name_l = name.lower()

    # Prospect signals: early interest, not enough info, intro email
    prospect_signals = [
        "shown interest",
        "not enough info",
        "intro email",
        "taking a guess",
        "has not been costed",
    ]
    if any(sig in text for sig in prospect_signals):
        return "prospect"

    # Costed signals: costing supplied to client
    if "costing supplied" in text or "wall costing supplied" in text:
        return "costed"

    # Committed signals: funding allocated, architect appointed, break ground, site visit with architect
    committed_signals = [
        "allocated funding",
        "break ground",
        "architect appointed",
        "site visit with architect",
    ]
    if any(sig in text for sig in committed_signals):
        return "committed"

    # Default: qualified (past initial contact but not yet costed/committed)
    return "qualified"


def read_excel(path: str) -> list[dict]:
    """Read Excel and return list of normalized project dicts."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Build header map (resilient to column reordering)
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    header_map = {}
    for i, h in enumerate(header_row):
        if h is not None:
            canonical = h.strip().lstrip(" ")  # Kill leading non-breaking space
            header_map[canonical] = i

    projects = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        name_raw = row[header_map["Project Name"]]
        if not name_raw or not str(name_raw).strip():
            continue

        name = str(name_raw).strip()
        pid = project_id(name)

        # Raw values for completion date (before normalization)
        completion_raw = row[header_map["Estimate Completion Date"]]
        completion_raw_str = str(completion_raw).strip() if completion_raw else None

        # Normalize completion: could be a date or a duration string
        completion_date = normalize_date(completion_raw)

        # Duration: parse from raw completion string if it's not a date
        duration_months = parse_duration_months(completion_raw_str)

        notes_raw = row[header_map["Notes"]]
        risks_raw = row[header_map["Risks"]]
        notes = normalize_string(notes_raw)
        risks = normalize_string(risks_raw)

        stage = infer_stage(name, notes_raw if isinstance(notes_raw, str) else str(notes_raw or ""),
                           risks_raw if isinstance(risks_raw, str) else str(risks_raw or ""))

        project = {
            "project_id": pid,
            "project_name": name,
            "phase": normalize_string(row[header_map["Project Category"]]),
            "updated": normalize_date(row[header_map["Updated"]]),
            "area": normalize_string(row[header_map["Area"]]),
            "total_project_value": normalize_value(row[header_map["Total Project Value"]]),
            "twx_project_value": normalize_value(row[header_map["TWX Project Value"]]),
            "projected_gp_30pct": normalize_value(row[header_map["Projected GP @30%"]]),
            "construction_company": normalize_string(row[header_map["Construction Company"]]),
            "client": normalize_string(row[header_map["Client"]]),
            "client_funds": normalize_string(row[header_map["Client Funds if known"]]),
            "notes": notes,
            "risks": risks,
            "estimated_takeoff_date": normalize_date(row[header_map["Estimated Takeoff date"]]),
            "estimated_completion_date": completion_date,
            "estimated_completion_raw": completion_raw_str,
            "estimated_duration_months": duration_months,
            "inferred_stage": stage,
        }
        projects.append(project)

    return projects


# ── SQLite Schema ──────────────────────────────────────────────────────

SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS pipeline_stages (
    stage_key TEXT PRIMARY KEY,
    label TEXT NOT NULL,
    sort_order INTEGER NOT NULL
);

CREATE TABLE IF NOT EXISTS projects (
    project_id TEXT PRIMARY KEY,
    project_name TEXT NOT NULL,
    phase TEXT,
    updated TEXT,
    area TEXT,
    total_project_value REAL,
    twx_project_value REAL,
    projected_gp_30pct REAL,
    construction_company TEXT,
    client TEXT,
    client_funds TEXT,
    notes TEXT,
    risks TEXT,
    estimated_takeoff_date TEXT,
    estimated_completion_date TEXT,
    estimated_completion_raw TEXT,
    estimated_duration_months REAL,
    inferred_stage TEXT REFERENCES pipeline_stages(stage_key),
    first_seen_date TEXT NOT NULL,
    last_seen_date TEXT NOT NULL,
    is_active INTEGER NOT NULL DEFAULT 1
);

CREATE TABLE IF NOT EXISTS project_history (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id TEXT NOT NULL REFERENCES projects(project_id),
    change_type TEXT NOT NULL,
    field_name TEXT,
    old_value TEXT,
    new_value TEXT,
    extraction_run_id INTEGER,
    changed_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS extraction_runs (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_at TEXT NOT NULL,
    source_file TEXT NOT NULL,
    source_checksum TEXT NOT NULL,
    projects_found INTEGER NOT NULL DEFAULT 0,
    projects_new INTEGER NOT NULL DEFAULT 0,
    projects_changed INTEGER NOT NULL DEFAULT 0,
    projects_removed INTEGER NOT NULL DEFAULT 0,
    forced INTEGER NOT NULL DEFAULT 0
);
"""

# Fields to track for change detection (excludes lifecycle fields)
TRACKED_FIELDS = [
    "project_name", "phase", "updated", "area",
    "total_project_value", "twx_project_value", "projected_gp_30pct",
    "construction_company", "client", "client_funds",
    "notes", "risks",
    "estimated_takeoff_date", "estimated_completion_date",
    "estimated_completion_raw", "estimated_duration_months",
    "inferred_stage",
]


def init_db(conn: sqlite3.Connection):
    conn.executescript(SCHEMA_SQL)
    # Seed pipeline_stages (idempotent via INSERT OR IGNORE)
    for key, label, order in STAGES:
        conn.execute(
            "INSERT OR IGNORE INTO pipeline_stages (stage_key, label, sort_order) VALUES (?, ?, ?)",
            (key, label, order),
        )
    conn.commit()


def get_existing_projects(conn: sqlite3.Connection) -> dict[str, dict]:
    """Return dict of project_id -> row dict for all projects."""
    cur = conn.execute("SELECT * FROM projects")
    cols = [desc[0] for desc in cur.description]
    result = {}
    for row in cur:
        d = dict(zip(cols, row))
        result[d["project_id"]] = d
    return result


def last_run_checksum(conn: sqlite3.Connection) -> str | None:
    row = conn.execute(
        "SELECT source_checksum FROM extraction_runs ORDER BY id DESC LIMIT 1"
    ).fetchone()
    return row[0] if row else None


def diff_and_write(conn: sqlite3.Connection, new_projects: list[dict], run_id: int, now: str):
    """Compare new_projects against existing DB, upsert projects, log changes."""
    existing = get_existing_projects(conn)
    existing_ids = set(existing.keys())
    new_ids = set(p["project_id"] for p in new_projects)

    stats = {"new": 0, "changed": 0, "removed": 0}

    for proj in new_projects:
        pid = proj["project_id"]

        if pid not in existing_ids:
            # New project
            conn.execute(
                """INSERT INTO projects (
                    project_id, project_name, phase, updated, area,
                    total_project_value, twx_project_value, projected_gp_30pct,
                    construction_company, client, client_funds,
                    notes, risks,
                    estimated_takeoff_date, estimated_completion_date,
                    estimated_completion_raw, estimated_duration_months,
                    inferred_stage, first_seen_date, last_seen_date, is_active
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (
                    pid, proj["project_name"], proj["phase"], proj["updated"], proj["area"],
                    proj["total_project_value"], proj["twx_project_value"], proj["projected_gp_30pct"],
                    proj["construction_company"], proj["client"], proj["client_funds"],
                    proj["notes"], proj["risks"],
                    proj["estimated_takeoff_date"], proj["estimated_completion_date"],
                    proj["estimated_completion_raw"], proj["estimated_duration_months"],
                    proj["inferred_stage"], now, now, 1,
                ),
            )
            conn.execute(
                "INSERT INTO project_history (project_id, change_type, extraction_run_id, changed_at) VALUES (?, 'new_project', ?, ?)",
                (pid, run_id, now),
            )
            stats["new"] += 1
        else:
            # Existing project – check for changes
            old = existing[pid]
            changed_fields = []
            for field in TRACKED_FIELDS:
                old_val = old.get(field)
                new_val = proj.get(field)
                # Normalize for comparison (both to string or both None)
                old_cmp = str(old_val) if old_val is not None else None
                new_cmp = str(new_val) if new_val is not None else None
                if old_cmp != new_cmp:
                    changed_fields.append((field, old_val, new_val))

            if changed_fields:
                # Update the project row
                set_clause = ", ".join(f"{f} = ?" for f in TRACKED_FIELDS)
                vals = [proj.get(f) for f in TRACKED_FIELDS]
                vals.extend([now, 1, pid])  # last_seen_date, is_active, WHERE project_id
                conn.execute(
                    f"UPDATE projects SET {set_clause}, last_seen_date = ?, is_active = ? WHERE project_id = ?",
                    vals,
                )
                for field, old_val, new_val in changed_fields:
                    change_type = "stage_change" if field == "inferred_stage" else "value_change"
                    conn.execute(
                        """INSERT INTO project_history
                           (project_id, change_type, field_name, old_value, new_value, extraction_run_id, changed_at)
                           VALUES (?, ?, ?, ?, ?, ?, ?)""",
                        (pid, change_type, field, str(old_val) if old_val is not None else None,
                         str(new_val) if new_val is not None else None, run_id, now),
                    )
                stats["changed"] += 1
            else:
                # No changes, just update last_seen
                conn.execute(
                    "UPDATE projects SET last_seen_date = ?, is_active = ? WHERE project_id = ?",
                    (now, 1, pid),
                )

            # Reappeared?
            if old.get("is_active") == 0:
                conn.execute(
                    "INSERT INTO project_history (project_id, change_type, extraction_run_id, changed_at) VALUES (?, 'reappeared', ?, ?)",
                    (pid, run_id, now),
                )

    # Removed projects (in DB but not in new extraction)
    removed_ids = existing_ids - new_ids
    for pid in removed_ids:
        if existing[pid].get("is_active") == 1:
            conn.execute("UPDATE projects SET is_active = 0 WHERE project_id = ?", (pid,))
            conn.execute(
                "INSERT INTO project_history (project_id, change_type, extraction_run_id, changed_at) VALUES (?, 'removed', ?, ?)",
                (pid, run_id, now),
            )
            stats["removed"] += 1

    return stats


def export_json(conn: sqlite3.Connection, now: str):
    """Export pipeline data to JSON for Streamlit consumption."""
    # Projects
    cur = conn.execute(
        "SELECT * FROM projects WHERE is_active = 1 ORDER BY phase, inferred_stage, project_name"
    )
    cols = [desc[0] for desc in cur.description]
    projects = [dict(zip(cols, row)) for row in cur]

    # Totals
    total_twx = sum(p["twx_project_value"] or 0 for p in projects)
    total_project = sum(p["total_project_value"] or 0 for p in projects)
    total_gp = sum(p["projected_gp_30pct"] or 0 for p in projects)

    # By stage
    stage_rows = conn.execute(
        "SELECT stage_key, label, sort_order FROM pipeline_stages ORDER BY sort_order"
    ).fetchall()
    by_stage = []
    for key, label, order in stage_rows:
        stage_projects = [p for p in projects if p["inferred_stage"] == key]
        by_stage.append({
            "key": key,
            "label": label,
            "sort_order": order,
            "count": len(stage_projects),
            "total_twx_zar": sum(p["twx_project_value"] or 0 for p in stage_projects),
            "total_project_zar": sum(p["total_project_value"] or 0 for p in stage_projects),
        })

    # By phase
    phases = sorted(set(p["phase"] for p in projects if p["phase"]))
    by_phase = {}
    for phase in phases:
        phase_projects = [p for p in projects if p["phase"] == phase]
        by_phase[phase] = {
            "count": len(phase_projects),
            "total_twx_zar": sum(p["twx_project_value"] or 0 for p in phase_projects),
            "total_project_zar": sum(p["total_project_value"] or 0 for p in phase_projects),
        }

    # By area
    areas = sorted(set(p["area"] for p in projects if p["area"]))
    by_area = {}
    for area in areas:
        area_projects = [p for p in projects if p["area"] == area]
        by_area[area] = {
            "count": len(area_projects),
            "total_twx_zar": sum(p["twx_project_value"] or 0 for p in area_projects),
        }

    # Recent changes (last 50)
    changes = conn.execute(
        """SELECT ph.changed_at, p.project_name, ph.change_type, ph.field_name, ph.old_value, ph.new_value
           FROM project_history ph
           JOIN projects p ON ph.project_id = p.project_id
           ORDER BY ph.id DESC LIMIT 50"""
    ).fetchall()
    recent_changes = []
    for changed_at, proj_name, change_type, field, old_val, new_val in changes:
        detail = ""
        if change_type == "new_project":
            detail = "Added to pipeline"
        elif change_type == "removed":
            detail = "Removed from pipeline"
        elif change_type == "reappeared":
            detail = "Reappeared in pipeline"
        elif field:
            detail = f"{field}: {old_val} -> {new_val}"
        recent_changes.append({
            "date": changed_at,
            "project": proj_name,
            "type": change_type,
            "detail": detail,
        })

    # Clean project dicts for JSON (convert booleans)
    clean_projects = []
    for p in projects:
        cp = dict(p)
        cp["is_active"] = bool(cp.get("is_active"))
        clean_projects.append(cp)

    data = {
        "_metadata": {
            "last_extraction": now,
            "total_projects": len(projects),
            "total_twx_value_zar": total_twx,
            "total_project_value_zar": total_project,
            "total_gp_30pct_zar": total_gp,
        },
        "pipeline_stages": by_stage,
        "projects": clean_projects,
        "summary": {
            "by_phase": by_phase,
            "by_stage": {s["key"]: {"count": s["count"], "total_twx_zar": s["total_twx_zar"]} for s in by_stage},
            "by_area": by_area,
        },
        "recent_changes": recent_changes,
    }

    os.makedirs(os.path.dirname(JSON_PATH), exist_ok=True)
    with open(JSON_PATH, "w") as f:
        json.dump(data, f, indent=2, default=str)

    return data


def main():
    parser = argparse.ArgumentParser(description="Extract TWX Sales Pipeline to SQLite + JSON")
    parser.add_argument("--force", action="store_true", help="Force extraction even if checksum unchanged")
    args = parser.parse_args()

    # Verify source exists
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Source file not found: {EXCEL_PATH}")
        sys.exit(1)

    now = datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S")
    checksum = file_checksum(EXCEL_PATH)
    print(f"Source: {os.path.basename(EXCEL_PATH)}")
    print(f"Checksum: {checksum[:16]}...")

    # Init DB
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA journal_mode=WAL")
    init_db(conn)

    # Checksum gate
    if not args.force and last_run_checksum(conn) == checksum:
        print("\nNo changes detected (source file checksum matches last run).")
        print("Use --force to re-extract anyway.")
        conn.close()
        return

    # Read and normalize
    print("\nReading Excel...")
    projects = read_excel(EXCEL_PATH)
    print(f"Found {len(projects)} projects")

    # Log extraction run
    cur = conn.execute(
        "INSERT INTO extraction_runs (run_at, source_file, source_checksum, forced) VALUES (?, ?, ?, ?)",
        (now, os.path.basename(EXCEL_PATH), checksum, int(args.force)),
    )
    run_id = cur.lastrowid

    # Diff and write
    stats = diff_and_write(conn, projects, run_id, now)

    # Update run stats
    conn.execute(
        """UPDATE extraction_runs SET projects_found = ?, projects_new = ?,
           projects_changed = ?, projects_removed = ? WHERE id = ?""",
        (len(projects), stats["new"], stats["changed"], stats["removed"], run_id),
    )
    conn.commit()

    # Export JSON
    data = export_json(conn, now)
    conn.close()

    # Print summary
    print(f"\n{'='*60}")
    print(f"TWX Pipeline Extraction Summary")
    print(f"{'='*60}")
    print(f"  Projects found:   {len(projects)}")
    print(f"  New:              {stats['new']}")
    print(f"  Changed:          {stats['changed']}")
    print(f"  Removed:          {stats['removed']}")
    print()

    # Phase breakdown
    by_phase = data["summary"]["by_phase"]
    for phase, info in sorted(by_phase.items()):
        print(f"  {phase}: {info['count']} projects, TWX value R {info['total_twx_zar']:,.0f}")

    # Stage breakdown
    print()
    for stage in data["pipeline_stages"]:
        if stage["count"] > 0:
            print(f"  {stage['label']:12s}: {stage['count']} projects, TWX R {stage['total_twx_zar']:,.0f}")

    print()
    meta = data["_metadata"]
    print(f"  Total TWX value:     R {meta['total_twx_value_zar']:,.0f}")
    print(f"  Total project value: R {meta['total_project_value_zar']:,.0f}")
    print(f"  Projected GP @30%:   R {meta['total_gp_30pct_zar']:,.0f}")
    print()
    print(f"  Database: {DB_PATH}")
    print(f"  JSON:     {JSON_PATH}")
    print(f"  DB size:  {os.path.getsize(DB_PATH):,} bytes")


if __name__ == "__main__":
    main()
